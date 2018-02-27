# Module with classes and functions related to CV information

# openpyxl helps to work with MSExcel databases
import openpyxl

# We import the following to create plots
import numpy as np
import matplotlib.pyplot as plt

# SUPORTED CLASSES
# CVItem
# -- Certificate
# ---- Award
# ---- Education
# ---- Job
# -- Event
# ---- Contest
# ---- Course
# ---- Talk
# ---- ScienceFair
# -- PortfolioItem
# ---- ArtItem
# ---- CodingItem
# -- Publication
# ---- Book
# ---- MainstreamMedia
# ---- ScientificPublication

# Level 0 Items

class CVItem(object):
    def __init__(self, *, id, title, date, description, links, tags):
        self.id = id
        self.title = title
        self.date = date
        self.description = description
        self.links = links
        self.tags = tags

# Level 1 Items

class Certificate(CVItem):
    def __init__(self, *, institution, location, **kwargs):
        super().__init__(**kwargs)
        self.institution = institution
        self.location = location

class Event(CVItem):
    def __init__(self, *, location, role, **kwargs):
        super().__init__(**kwargs)
        self.location = location
        self.role = role

class PortfolioItem(CVItem):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class Publication(CVItem):
    def __init__(self, *, authors, publisher, **kwargs):
        super().__init__(**kwargs)
        self.authors = authors
        self.publisher = publisher

# Level 2 Items

class Award(Certificate):
    sheet="Award"
    def __init__(self, **kwargs):
        super().__init__(links=None, **kwargs)

class Education(Certificate):
    sheet="Education"
    def __init__(self, *, start_date, end_date, **kwargs):
        super().__init__(date=[start_date,end_date], links=None, **kwargs)

class Job(Certificate):
    sheet="Job"
    def __init__(self, *, start_date, end_date, **kwargs):
        super().__init__(date=[start_date,end_date], links=None, **kwargs)

class Contest(Event):
    sheet="Contest"
    def __init__(self, *, start_date, end_date,audience, **kwargs):
        super().__init__(links=None, date=[start_date,end_date], **kwargs)
        self.audience=audience

class Course(Event):
    sheet="Course"
    def __init__(self, *, institution, **kwargs):
        super().__init__(description="", links=None, **kwargs)
        self.institution=institution

class Talk(Event):
    sheet="Talk"
    def __init__(self, *, audience, institution, file, **kwargs):
        super().__init__(role="lecturer", links={"file":file}, **kwargs)
        self.audience=audience
        self.institution=institution

class ScienceFair(Event):
    sheet="ScienceFair"
    def __init__(self, *, audience, institution, **kwargs):
        super().__init__(links=None, **kwargs)
        self.audience=audience
        self.institution=institution

class ArtItem(PortfolioItem):
    sheet="ArtItem"
    def __init__(self, *, technique, youtube, file, **kwargs):
        super().__init__(links={"file":file, "youtube":youtube}, **kwargs)
        self.technique=technique

class CodingItem(PortfolioItem):
    sheet="CodingItem"
    def __init__(self, *, technologies, github, web, **kwargs):
        super().__init__(links={"github":github, "web":web},**kwargs)
        self.technologies=technologies

class Book(Publication):
    sheet="Book"
    def __init__(self, *, file, **kwargs):
        super().__init__(links={"file":file}, **kwargs)

class MainstreamMedia(Publication):
    sheet="MainstreamMedia"
    def __init__(self, *, media_name, media_type, file, **kwargs):
        super().__init__(links={"file":file}, **kwargs)
        self.media_name=media_name
        self.media_type=media_type

class ScientificPublication(Publication):
    sheet="ScientificPublication"
    def __init__(self, *, journal, volume, pages, doi, arxiv, web, **kwargs):
        super().__init__(links={"arxiv":arxiv, "doi":doi, "web":web}, **kwargs)
        self.journal=journal
        self.volume=volume
        self.pages=pages



### FUNCTIONS ###
# Functions to load the items from the CV

def fetch_CV(location):
    return openpyxl.load_workbook(location, data_only=True)

def fetch_from_CV(cv,class_,querytags=None):
    data_ws = cv[class_.sheet]
    params = {}
    items = []
    y = 3 #y compensates for column titles
    while data_ws.cell(row=y,column=1).value is not None:
        x = 1 
        while data_ws.cell(row=1,column=x).value is not None:
            param = data_ws.cell(row=2,column=x).value
            value = data_ws.cell(row=y,column=x).value
            params[param] = value
            x+=1
        if querytags is None:
            new_item=class_(**params)
            items+=[new_item]
        else:
            for tag in querytags:
                if tag in params["tags"]:
                    new_item=class_(**params)
                    items+=[new_item]
                    break
        y+=1
    return items

if __name__=="__main__":
    cv = fetch_CV(r'data\CV.xlsx')
    education_array = fetch_from_CV(cv,Education)
    papers_array = fetch_from_CV(cv, ScientificPublication, querytags=["thesis", "journal"])
    for e in education_array:
        print(e.title)
    print()
    for p in papers_array:
        print(p.title)
