# openpyxl helps to work with MSExcel databases
import openpyxl as opx

# We import the following to create plots
import numpy as npy
import matplotlib.pyplot as plt

# class ItemList(object):
#     # This is just a collection (dict) of items . We may use it to collect all the 
#     # items in a CV, or in Sections or a filtered collection of them.

#     def __init__(self, *, id, name, title, items=[]):
#         self.id=id
#         self.name=name
#         self.title=title
#         self.items=items

#     def filter_by_tag(tags):
#         pass


class Section(object):
    # We think of sections as "filters".
    #
    # Here is where OO kicks in. For a better management, we want to group
    # together items that are similar. We handle this as thinking them of
    # objects with the same properties, so all the items in a Section must
    # be of the same CVItem subclass.
    # 
    # Moreover, each section needs a list of filter tags.
    # We enforce that for each class a Section either has all or none of the
    # instances of its corresponding class with that tag.
    #
    # This is way better for organizing front end CVs.
    
    def __init__(self, *, id, name, title, _class, filter_tags):
        self.id = id
        self.name = name
        self.title = title
        self._class=_class # A subclass of CVItem
        if type(filter_tags) is str:
            self.filter_tags=filter_tags.split(", ")
        else:
            self.filter_tags=filter_tags
        if self.filter_tags == None:
            self.filter_tags_set=set()
        else:
            self.filter_tags_set=set(self.filter_tags)

class CurriculumVitae(object):
    # A CV is a collection of CVItems organized by CVItem subclasses
    # This is implemented via a dictionary of the subclasses. Each
    # key (i.e. a CVItem subclass) yields the list of elements corresponding
    # to that class.
    # It is also a collection of sections. They are handled by a dict
    # too, using the name of the section as the key.

    def __init__(self, *, id_name, all_items, sections):
        self.id_name=id_name
        self.all_items=all_items
        self.sections=sections
        # self.link_handler=link_handler

    def item_classes(self):
        print(self.all_items.keys())

    def fetch_by_tag(self, filter_tags):
        result=[]
        if type(filter_tags) is str:
            filter_tags = filter_tags.split(", ")
        filter_tag_set=set(filter_tags)
        for _, item_class_list in self.all_items:
            for item in item_class_list:
                if bool(filter_tag_set & item.tags_set):
                    result+=[item]
        return result

    def fetch_by_class(self, _class):
        return self.all_items(_class)

    def fetch_by_section(self, section):
        result=[]
        items_class_list = self.all_items[section._class]
        filter_tag_set = section.filter_tags_set
        for item in items_class_list:
            if bool(filter_tag_set & item.tags_set):
                result += [item]
        return result

    def fetch_by_section_name(self,name):
        return self.fetch_by_section(self.sections[name])


### Functions ###

# def sec_from_xlsx(location,worksheet,filters)
#     CV_ws = cv_from_xlsx(location)

def cv_from_xlsx(location):
    cv_db = opx.load_workbook(location, data_only=True)

    section_ws = cv_db["SECTION"]    
    seclist_ws = cv_db["SECLIST"]
    config_ws= cv_db["CONFIG"]
    link_ws= cv_db["LINK"]

    # The following line is needed for  str -> class
    class_lib = __import__('cvutils.primaryclasses')

    id_name = config_ws.cell(row=2,column=1).value
    item_classes_names=[]
    item_classes=[]
    all_items={}
    sections={}

    # Fetch classes names in a list of strings
    y=5 # headers offset
    while config_ws.cell(row=y,column=1).value is not None:
        if config_ws.cell(row=y, column=2).value=="True":
            # Note the conversion str -> class
            item_classes_names+=[config_ws.cell(row=y, column=1).value]
        y+=1

    # Fetch sections
    params= [section_ws.cell(row=2, column=x).value for x in range(1,6)]
    
    y=3 
    while section_ws.cell(row=y,column=1).value is not None:
        values = [section_ws.cell(row=y, column=x).value for x in range(1,6)]
        # str -> class
        values[3] = getattr(class_lib, values[3])
        params_dict=dict(zip(params,values))
        section=Section(**params_dict)
        sections[section.name]=section
        y+=1

    # Fetch all items and create dictionary of item classes

    for class_name in item_classes_names:
        item_class_ws=cv_db[class_name]

        # This deals with the list of item classes
        _class = getattr(class_lib, class_name)
        class_items=[]

        # The fetching
        x=1
        params=[]
        while item_class_ws.cell(row=2, column=x).value is not None:
            param=item_class_ws.cell(row=2, column=x).value
            params += [param]
            x+=1
        num_params=x-1
        y=3
        while item_class_ws.cell(row=y,column=1).value is not None:
            values=[]
            for z in range(num_params):
                values+=[item_class_ws.cell(row=y,column=z+1).value]
            params_dict = dict(zip(params, values))
            new_item = _class(**params_dict)
            class_items+=[new_item]
            y+=1

        all_items[_class]=class_items

    return CurriculumVitae(id_name=id_name, all_items=all_items, sections=sections)

def fetch_from_CV(cv,class_,querytags=None):
    data_ws = cv[class_.sheet]
    params = {}
    items = []
    y = 3 #y compensates for column titles

    while data_ws.cell(row=y,column=1).value is not None:
        x = 1
        while data_ws.cell(row=1,column=x).value is not None:
            param = data_ws.cell(row=2,column=x).value
            value = data_ws.cell(row=y,column=x).value
            params[param] = value
            x+=1
        if querytags is None:
            new_item=class_(**params)
            items+=[new_item]
        else:
            for tag in querytags:
                if tag in params["tags"]:
                    new_item=class_(**params)
                    items+=[new_item]
                    break
        y+=1
    return items
