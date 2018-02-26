# Module with classes and functions related to CV information

# openpyxl helps to work with MSExcel databases
import openpyxl

# We import the following to create plots
import numpy as np
import matplotlib.pyplot as plt

# CLASSES
#

class ArtItem:
    num_attr = 6
    sheet = "Art"
    def __init__(self, kind, title, year, technique, link, description):
        self.kind=kind
        self.title=title
        self.year=year
        self.technique=technique
        self.description=description
        self.link=link


class Award:
    num_attr = 5
    sheet = "Awards"
    def __init__(self, year, award, institution, description, location):
        self.year=year
        self.award=award
        self.institution=institution
        self.description=description
        self.location=location


class Job:
    num_attr = 4
    sheet = "Jobs"
    def __init__(self, position, period, employer, description):
        self.position=position
        self.period=period
        self.employer=employer
        self.description=description


class Paper:
    num_attr = 8
    sheet = "Papers"
    def __init__(self, kind, title, authors,
         journal, year, volume, pages, doi):
        self.kind= kind
        self.title=title
        self.authors=authors
        self.journal=journal
        self.year=year
        self.volume=volume
        self.pages=pages
        self.doi=doi


class Presentation:
    num_attr = 7
    sheet = "Presentations"
    def __init__(self, kind, title, venue, conference,
         year, link):
        self.kind=kind
        self.title=title
        self.venue=venue
        self.conference=conference
        self.year=year
        self.link=link
        self.location=location


class ProgrammingItem:
    num_attr = 6
    sheet = "Programming"
    def __init__(self, title, year, kind, technologies, description, link):
        self.title=title
        self.year=year
        self.kind=kind
        self.technologies=technologies
        self.description=description
        self.link=link


### FUNCTIONS ###
# Functions to load the items from the CV

def fetch_CV(location):
    return openpyxl.load_workbook(location, data_only=True)

def fetch_from_CV(cv,class_):
    data_ws = cv[class_.sheet]
    item_array = []
    num_attr = class_.num_attr
    j = 3 #Compensate for DB column titles
    while data_ws.cell(row=j,column=1).value is not None:
        new_item=class_(
            *[data_ws.cell(row=j,column=k).value for k in range(2,2+num_attr)]
            )
        item_array+=[new_item]
        j+=1
    return item_array

def papers_timeline(cv,xkcd=False):
    paper_array = fetch_from_CV(cv,Paper)
    years=[paper.year for paper in paper_array]
    values=[y for y in range(min(years),max(years)+1)]
    count=[years.count(value) for value in values]
    labels=tuple([str(value) for value in values])

    y_pos = np.arange(len(values))
    plt.rcdefaults()
    for k in ["axes.edgecolor", "axes.labelcolor", "xtick.color", "ytick.color"]:
        plt.rcParams[k]=[0,0,43/255]
    plt.rcParams["figure.figsize"]=[8,2]
    plt.axes()

    if xkcd is True:
        plt.xkcd() #Fun!

    plt.bar(y_pos, count, align='center', width=0.4, linewidth=0,color=[200/255,55/255,113/255])
    plt.xticks(y_pos, labels)
    plt.yticks(list(range(0,max(count)+2)))
    plt.ylabel('Publications')
    plt.savefig(r'static/papers.png', transparent=True)
    plt.clf()

def python_timeline(cv,xkcd=False):
    data_array = fetch_from_CV(cv,ProgrammingItem)
    years=[item.year for item in data_array]
    values=[y for y in range(min(years),max(years)+1)]
    count=[years.count(value) for value in values]
    labels=tuple([str(value) for value in values])

    y_pos = np.arange(len(values))
    plt.rcdefaults()
    for k in ["axes.edgecolor", "axes.labelcolor", "xtick.color", "ytick.color"]:
        plt.rcParams[k]=[0,0,43/255]
    plt.rcParams["figure.figsize"]=[8,2]
    plt.axes()

    if xkcd is True:
        plt.xkcd() #Fun!
        
    plt.bar(y_pos, count, align='center', width=0.4, linewidth=0,color=[200/255,55/255,113/255])
    plt.xticks(y_pos, labels)
    plt.yticks(list(range(0,max(count)+2)))
    plt.ylabel('Projects')
    plt.savefig(r'static/programming.png', transparent=True)
    plt.clf()
    print("Exito")